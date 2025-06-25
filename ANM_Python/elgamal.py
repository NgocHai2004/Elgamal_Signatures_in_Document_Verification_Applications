import hashlib
import random
import docx
import PyPDF2
import base64
import os
from enum import Enum
import base64
import binascii
import openpyxl

HASH_ALGORITHMS = {
    "MD5": hashlib.md5,
    "SHA-1": hashlib.sha1,
    "SHA-256": hashlib.sha256,
    "SHA-384": hashlib.sha384,
    "SHA-512": hashlib.sha512,
}

class VerificationStatus(Enum):
    VALID = 0
    CRYPTO_MISMATCH = 3

class SignatureDecodeError(ValueError):
    def __init__(self, messages):
        self.messages = messages
        super().__init__("\n- ".join(["Lỗi định dạng chữ ký:"] + messages))

def read_file_bytes(file_path: str) -> bytes:
    with open(file_path, "rb") as f:
        return f.read()

def get_displayable_text_from_file(file_path: str) -> str:
    data_list = []
    filename = os.path.basename(file_path)
    file_extension = os.path.splitext(file_path)[1].lower()
    
    try:
        if not os.path.exists(file_path):
            return f"[File not found: {filename}]"

        if file_extension == ".docx":
            try:
                doc = docx.Document(file_path)
                for para in doc.paragraphs: data_list.append(para.text)
                data = "\n".join(filter(None, data_list))
                if not data.strip() and doc.element.body:
                    data = f"[DOCX '{filename}' - No text extracted, possibly image-only or complex]"
            except Exception as e: return f"[Error reading DOCX '{filename}': {e}]"

        elif file_extension == ".pdf":
            try:
                with open(file_path, "rb") as pdf_file_obj:
                    pdf_reader = PyPDF2.PdfReader(pdf_file_obj)
                    if pdf_reader.is_encrypted:
                        password_types = getattr(PyPDF2, 'PasswordType', None)
                        decrypt_result = None
                        try: decrypt_result = pdf_reader.decrypt('')
                        except Exception: 
                             if pdf_reader.is_encrypted: return f"[Encrypted PDF '{filename}' - Decryption failed/needs password]"
                        
                        is_still_encrypted = True
                        if password_types:
                            if decrypt_result == password_types.OWNER_PASSWORD or decrypt_result == password_types.USER_PASSWORD:
                                is_still_encrypted = False
                        elif decrypt_result :
                            is_still_encrypted = False
                        
                        if is_still_encrypted and pdf_reader.is_encrypted :
                             return f"[Encrypted PDF '{filename}' - Cannot extract (password needed)]"
                    
                    for i, page in enumerate(pdf_reader.pages):
                        try:
                            page_text = page.extract_text()
                            if page_text: data_list.append(page_text)
                        except Exception as e_page: data_list.append(f"[Page {i+1} in '{filename}': Error - {e_page}]")
                    
                    data = "\n".join(filter(None, data_list))
                    if not data.strip() and len(pdf_reader.pages) > 0:
                        data = f"[PDF '{filename}' ({len(pdf_reader.pages)} pages) - No text extracted or image-based]"
            except PyPDF2.errors.PdfReadError as e: return f"[Error PDF '{filename}': {e} - Corrupted/Not PDF?]"
            except Exception as e: return f"[Could not process PDF '{filename}': {e}]"
            
        elif file_extension in [".xlsx", ".xls"]:
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                data_list.append(str(cell.value))
                data = "\n".join(filter(None, data_list))
                if not data.strip() and len(wb.sheetnames) > 0:
                     data = f"[Excel '{filename}' ({len(wb.sheetnames)} sheets) - No data extracted]"
            except Exception as e:
                return f"[Error reading Excel file '{filename}': {e}]"

        else: 
            try:
                with open(file_path, "r", encoding="utf-8", errors="replace") as f: data = f.read()
                if not data.strip() and os.path.getsize(file_path) > 0:
                    common_binary_exts = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.exe', '.dll', '.zip', '.gz', '.rar', '.mp3', '.mp4', '.avi', '.mov', '.dat', '.bin', '.iso', '.img', '.doc', '.ppt'}
                    if file_extension in common_binary_exts:
                        data = f"[Binary file ({file_extension}) '{filename}' - Not displayable]"
                    else: data = f"[File '{filename}' - No displayable text (binary or non-UTF-8?)]"
            except UnicodeDecodeError: data = f"[Binary/non-UTF-8 '{filename}' - Cannot display]"
            except Exception as e: return f"[Error reading text file '{filename}': {e}]"
    except Exception as e: return f"[Error processing file '{filename}': {e}]"
    
    if not data.strip():
        if os.path.exists(file_path) and os.path.getsize(file_path) == 0: return f"[File '{filename}' is empty]"
        if not data.startswith("["): return f"[No textual content from '{filename}']"
            
    return data.strip()

def hash_data_to_int(data_bytes: bytes, mode: str) -> int:
    hasher = HASH_ALGORITHMS.get(mode)
    if not hasher: raise ValueError(f"Unsupported hash mode: {mode}")
    return int(hasher(data_bytes).hexdigest(), 16)

def encode_signature(r: int, s: int, format_type: str) -> str:
    if format_type.lower() not in ["hex", "base64"]:
        raise ValueError("Format must be 'hex' or 'base64'")
    
    r_bl = r.bit_length(); s_bl = s.bit_length()
    r_bytes = r.to_bytes((r_bl + 7) // 8 if r_bl > 0 else 1, byteorder='big')
    s_bytes = s.to_bytes((s_bl + 7) // 8 if s_bl > 0 else 1, byteorder='big')
    
    if format_type.lower() == "hex":
        return f"{r_bytes.hex()};{s_bytes.hex()}"
    else:
        r_b64 = base64.urlsafe_b64encode(r_bytes).decode('utf-8')
        s_b64 = base64.urlsafe_b64encode(s_bytes).decode('utf-8')
        return f"{r_b64};{s_b64}"

def decode_signature(signature_str: str, format_type: str) -> tuple[int, int]:
    if format_type.lower() not in ["hex", "base64"]:
        raise ValueError("Định dạng phải là 'hex' hoặc 'base64'")

    cleaned_str = signature_str.strip()
    parts = cleaned_str.split(';')
    if len(parts) != 2:
        raise SignatureDecodeError([f"Định dạng chữ ký không đúng. Cần có dạng 'phần_r;phần_s' nhưng nhận được {len(parts)} phần."])

    r_part_str, s_part_str = parts[0], parts[1]
    error_messages = []
    r, s = None, None

    if not r_part_str:
        error_messages.append("Phần 'r' của chữ ký bị trống.")
    else:
        try:
            if format_type.lower() == "hex":
                r_bytes_decoded = bytes.fromhex(r_part_str)
                r = int.from_bytes(r_bytes_decoded, byteorder='big')
            else: 
                r_bytes_decoded = base64.urlsafe_b64decode(r_part_str)
                r = int.from_bytes(r_bytes_decoded, byteorder='big')
                
                r_part_re_encoded_b64 = base64.urlsafe_b64encode(r_bytes_decoded).decode('utf-8')
                if r_part_re_encoded_b64 != r_part_str:
                    raise ValueError("Base64 for 'r' has issues.") 
        except (ValueError, base64.binascii.Error): 
            error_messages.append(f"Phần 'r' không hợp lệ (kiểm tra định dạng {format_type.upper()}).")

    if not s_part_str:
        error_messages.append("Phần 's' của chữ ký bị trống.")
    else:
        try:
            if format_type.lower() == "hex":
                s_bytes_decoded = bytes.fromhex(s_part_str)
                s = int.from_bytes(s_bytes_decoded, byteorder='big')
            else:
                s_bytes_decoded = base64.urlsafe_b64decode(s_part_str)
                s = int.from_bytes(s_bytes_decoded, byteorder='big')

                s_part_re_encoded_b64 = base64.urlsafe_b64encode(s_bytes_decoded).decode('utf-8')
                if s_part_re_encoded_b64 != s_part_str:
                    raise ValueError("Base64 for 's' has issues.")
        except (ValueError, base64.binascii.Error): 
            error_messages.append(f"Phần 's' không hợp lệ (kiểm tra định dạng {format_type.upper()}).")

    if error_messages:
        raise SignatureDecodeError(error_messages)

    return r, s

def find_gcd(a: int, b: int) -> int:
    while b: a, b = b, a % b
    return abs(a)

def is_probably_prime(n: int, k: int = 20) -> bool:
    if n <= 1: return False;
    if n <= 3: return True
    if n % 2 == 0: return False
    s, d = 0, n - 1
    while d % 2 == 0: s += 1; d //= 2
    for _ in range(k):
        a = random.randrange(2, n - 1)
        x = pow_mod(a, d, n)
        if x == 1 or x == n - 1: continue
        for _ in range(s - 1):
            x = pow_mod(x, 2, n)
            if x == n - 1: break
        else: return False
    return True

def check_prime(n: int) -> bool:
    if n < 2: return False
    if n < 1000:
        for i in range(2, int(n**0.5) + 1):
            if n % i == 0: return False
        return True
    return is_probably_prime(n, k=20)

def find_prime(bits: int) -> int:
    if bits < 8: raise ValueError("Bit size must be at least 8.")
    lower, upper = 1 << (bits - 1), (1 << bits) - 1
    for _ in range(2000):
        res = random.randrange(lower, upper + 1)
        if res % 2 == 0: res += 1
        if res > upper: continue
        if check_prime(res): return res
    raise ValueError(f"Cannot find {bits}-bit prime after many attempts.")

def pow_mod(a: int, b: int, n: int) -> int:
    res = 1; a %= n
    while b > 0:
        if b % 2 == 1: res = (res * a) % n
        a = (a * a) % n; b //= 2
    return res

def mod_inverse(a: int, m: int) -> int:
    if m == 0: raise ValueError("Modulus cannot be zero.")
    gcd_val = find_gcd(a, m)
    if gcd_val != 1: raise ValueError(f"Modular inverse does not exist (gcd({a},{m})={gcd_val} != 1).")
    m0, x0, x1 = m, 0, 1
    if m == 1: return 0
    a_orig = a
    while a > 1:
        if m == 0: raise ValueError(f"Modulus became zero for a={a_orig}, m={m0}")
        q = a // m; m_temp = m; m = a % m; a = m_temp
        x0_temp = x0; x0 = x1 - q * x0; x1 = x0_temp
    if x1 < 0: x1 += m0
    return x1

def gen_g(p: int) -> int:
    if p <= 3: raise ValueError("p must be > 3 for g generation.")
    return random.randint(2, p - 1) 

def gen_x(p: int) -> int:
    if p <= 4: raise ValueError("p must be > 4 for x generation.")
    return random.randint(2, p - 2)

def gen_k_for_signature(p: int) -> int:
    if p <= 3 or p - 1 == 0: raise ValueError("p must be > 3 for k generation.")
    for _ in range(p):
        k = random.randint(1, p - 2)
        if find_gcd(k, p - 1) == 1: return k
    raise ValueError(f"Cannot find valid k (coprime to p-1={p-1}) for p={p}.")

def cre_key(bits: int) -> tuple[int, int, int, int]:
    p = find_prime(bits); g = gen_g(p)
    x = gen_x(p); y = pow_mod(g, x, p)
    return p, g, x, y

def create_sign(data_bytes: bytes, hash_mode: str, p: int, g: int, x: int, k_for_signing: int) -> tuple[int, int]:
    m_int = hash_data_to_int(data_bytes, hash_mode)
    if not (0 < k_for_signing < p - 1 and find_gcd(k_for_signing, p - 1) == 1):
        raise ValueError(f"k={k_for_signing} invalid for p={p}. Must be (0, p-1) & gcd(k, p-1)=1.")
    
    r_sig = pow_mod(g, k_for_signing, p)
    try: k_inv = mod_inverse(k_for_signing, p - 1)
    except ValueError as e: raise ValueError(f"Error k_inv for k={k_for_signing}, p-1={p - 1}: {e}")
    
    s_num = (m_int - (x * r_sig)) % (p - 1)
    s_val = (s_num * k_inv) % (p - 1)
    if s_val == 0: raise ValueError("s=0 generated. Choose different k and retry.")
    return r_sig, s_val

def verify_sign(
        data_bytes: bytes, hash_mode: str, signature_str: str, format_type: str,
        p: int, g: int, y: int
) -> VerificationStatus: 
    r_rcvd, s_rcvd = decode_signature(signature_str, format_type)
    
    error_messages_range = []
    if not (0 < r_rcvd < p):
        error_messages_range.append(f"Giá trị 'r' ({r_rcvd}) nằm ngoài phạm vi cho phép (0 < r < p={p}).")
    if not (0 < s_rcvd < p - 1):
        error_messages_range.append(f"Giá trị 's' ({s_rcvd}) nằm ngoài phạm vi cho phép (0 < s < p-1={p-1}).")

    if error_messages_range:
        raise SignatureDecodeError(error_messages_range) 

    m_int = hash_data_to_int(data_bytes, hash_mode)
    
    lhs = pow_mod(g, m_int, p)
    term1_rhs = pow_mod(y, r_rcvd, p)
    term2_rhs = pow_mod(r_rcvd, s_rcvd, p)
    rhs = (term1_rhs * term2_rhs) % p
    
    return VerificationStatus.VALID if lhs == rhs else VerificationStatus.CRYPTO_MISMATCH